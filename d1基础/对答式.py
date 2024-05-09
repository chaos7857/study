import requests
from openpyxl import load_workbook
import json
import os
# direction = "C:\\Users\\Cc\\Desktop\\"
# filename = "青年大学习2023"
# TODO 解决Json文件绝对路径的定义问题

direction = os.getcwd()
direction = os.path.abspath(direction)
try:   
    with open(direction+"\\date.json", "r") as f:
        data = json.load(f)
    filename = data[1]
except:
    filename = "青年大学习2023"
def edit_url(section_number):
    a = 241 + 2 * section_number
    url = "http://dxx.ahyouth.org.cn/api/peopleRankStage?table_name=reason_stage"+ str(a) \
    + "&level1=%E7%9B%B4%E5%B1%9E%E9%AB%98%E6%A0%A1&level2=%E5%90%88%E8%82%A5%E5%B7%A5%E4%B8%9A%E5%A4%A7%E5%AD%A6&level3=%E7%AE%A1%E7%90%86%E5%AD%A6%E9%99%A2%E5%9B%A2%E5%A7%94&level4=%E4%BF%A1%E6%81%AF%E7%AE%A1%E7%90%86%E4%B8%8E%E4%BF%A1%E6%81%AF%E7%B3%BB%E7%BB%9F%E4%B8%93%E4%B8%9A2022%E7%BA%A71%E7%8F%AD"
    return url



def get_done_name(section_number):
    try:
        response = requests.get(edit_url(section_number)).json()
    except Exception as e:
        print("这期还没有哦!")
        return e
    else:
        list_n = []
        for name in response['list']['list']:
            list_n.append(name['username'])
        return list_n
    

def get_standard(direction, filename):
    try:
        wb = load_workbook("%s%s.xlsx"%(direction, filename))
    except Exception as e:
        print("该路径下找不到这个文件！\n请重新设置！")
        return e
    else:
        standard = []
        sheet1 = wb["Sheet1"]
        for each_cell in sheet1['B']:
            if each_cell.value != None:
                standard.append(each_cell.value)
        wb.close()
        return standard


def catch_undone(list_diff, n, direction, filename):
    try:
        wb = load_workbook("%s%s.xlsx"%(direction, filename))  
    except Exception as e:
        print("该路径下找不到这个文件！\n请重新设置！")
        return e
    else:  
        sheet1 = wb["Sheet1"]
        sheet1.cell(1, n+4).value = "第%d期"%n
        i = 2
        while i<100:

            for diff in list_diff:
                if sheet1.cell(i, 2).value == diff:
                    sheet1.cell(i, n+4).value = "还没做"  
                    break

            i+=1
        wb.save("%s%s.xlsx"%(direction, filename))
        wb.close()
        print("SUCCESS_SIGNED!")


def change_road(direction, filename):
    print("your direction now:%s"%direction)    
    direction_ = input("enter the direction:")
    if len(direction_):
        direction = direction_
    print("successful to change direction to %s"%direction)
    print("your filename now:%s"%filename)
    filename_ = input("enter the filename:")
    if len(filename_):
        filename = filename_
    print("successful to change direction to %s"%filename)
    with open("date.json", "w") as f:
        json.dump([direction, filename], f)
    




if __name__ == "__main__":
    while 1:
        try:
            print("\t\tWelcome to use!")
            print("【0】退出\t【1】开始查询\t【2】设置\t【3】清屏")

            m = int(input("请输入操作："))
            if m == 0:
                break
            elif m == 1:
                n = int(input("请输入要查询的期数："))
                list_n = get_done_name(n)
                list_st = get_standard(direction, filename)
                list_diff = [y for y in list_st if not y in list_n]
                print("SUCCESSFUL_EXIT!")
                print('='*50)
                if len(list_diff)==0:
                    print("%d人已经全部完成！"%len(list_n))
                else:    
                    print("还有 %d 人没完成: \n\n\n"%len(list_diff))
                    for value in list_diff:
                        print(value)
                    print('-'*50)
                    print("注意核对当前完成的总人数是%d哦！"%len(list_n))
                print('='*50)
                flag = input("需要给他们打标记吗？(y/n)：")
                if flag == 'y':
                    catch_undone(list_diff, n, direction, filename)
            elif m == 2:
                change_road(direction, filename)
            else:
                print("really?")
        except Exception as e:
            print(e)
