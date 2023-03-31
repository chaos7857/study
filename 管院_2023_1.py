import requests
from openpyxl import load_workbook
import os
import json


def welcome():
    print("\t\tWelcome to use!!")
    print("【1】查询\t【2】设置\t【0】退出")


def search():
    n = int(input("请输入查询期数："))
    url = "http://dxx.ahyouth.org.cn/api/peopleRankStage?table_name=reason_stage"+str(241 + 2 * n) +\
      "&level1=%E7%9B%B4%E5%B1%9E%E9%AB%98%E6%A0%A1&level2=%E5%90%88%E8%82%A5%E5%B7%A5%E4%B8%9A%E5%A4%A7%E5%AD%A6&" \
      "level3=%E7%AE%A1%E7%90%86%E5%AD%A6%E9%99%A2%E5%9B%A2%E5%A7%94"
    try:
        info = requests.get(url).json()
    except Exception as e:
        print(e)
    else:
        for ggg in info['list']['list']:
            name.append(ggg['level4'])
            num.append(ggg['num'])
    
def edit(sheet):
    st_name = []
    for each_cell in sheet['B']:
        if each_cell.value is not None:
            st_name.append(each_cell.value)

    for j in range(1, len(st_name)+1):
        sheet.cell(j, 23).value = 0

    for i in range(len(name)):
        for j in range(1, len(st_name)+1):
            if st_name[j-1] == name[i]:
                sheet.cell(j, 23).value += num[i]
                do_name.append(name[i])
    sheet.cell(1, 23).value = "success"
    

def ma_sure():
    print("将对《%s》进行编辑，该操作不可撤回！！！"%xlsx_name)
    print("【1】确认\t【0】返回")
    s_cho = int(input("决定？："))
    return s_cho

def settings(xlsx_name):
    print("当前操作表格：%s"%xlsx_name)
    print("【1】修改\t【0】确认")
    j = int(input("I select:"))
    if j == 1:
        
        xlsx_name = input("输入目标表格名称：")
        with open(direction+"\\date.json", "w") as f:
            json.dump(xlsx_name, f)
        print("成功将当前操作表格改为：%s"%xlsx_name)

                
    elif j == 0:
        pass
    else:
        print("请别为难我=。=")

if __name__ == "__main__":
    while 1:
        welcome()
        cho = int(input("请输入要执行的操作："))

        name = []
        num = []
        li_sh = []
        do_name = []
        direction = os.path.abspath(os.getcwd())+'\\'
        
        try:   
            with open(direction+"\\date.json", "r") as f:
                xlsx_name = json.load(f)
        except Exception:
            xlsx_name = "管理学院-青年大学习·2023"

        if cho == 0:
            break
        elif cho == 1:
            i = ma_sure()
            if i == 1:
                search()
                try:
                    wb = load_workbook("%s%s.xlsx"%(direction, xlsx_name))
                except Exception as e:
                    print(e)
                else:
                    for sheet_n in wb:
                        edit(sheet_n)

                    undo_name = [x for x in name if x not in do_name]
                    print("undo_list:%d" % len(undo_name))

                    for names in undo_name:
                        print(names)
                    
                    wb.save("%s%s.xlsx"%(direction, xlsx_name))
                    wb.close()

            elif i == 0:
                continue
            else:
                print("请别为难我=。=")

        elif cho == 2:
            settings(xlsx_name)
        else:
            print("请别为难我=。=")
