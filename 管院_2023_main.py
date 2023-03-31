import requests
from openpyxl import load_workbook
import os
direction = os.path.abspath(os.getcwd())+'\\'
xlsx_name = "管理学院-青年大学习·2023"

url = "http://dxx.ahyouth.org.cn/api/peopleRankStage?table_name=reason_stage"+"249" \
      "&level1=%E7%9B%B4%E5%B1%9E%E9%AB%98%E6%A0%A1&level2=%E5%90%88%E8%82%A5%E5%B7%A5%E4%B8%9A%E5%A4%A7%E5%AD%A6&" \
      "level3=%E7%AE%A1%E7%90%86%E5%AD%A6%E9%99%A2%E5%9B%A2%E5%A7%94"
info = requests.get(url).json()

name = []
num = []
li_sh = []
do_name = []

for ggg in info['list']['list']:
    name.append(ggg['level4'])
    num.append(ggg['num'])


# 用以检查是否有重复的组织名?
# name.sort()
# i = 1
# while(i<len(name)):
#     if name[i] == name[i-1]:
#         print(name[i])
#     i+=1
# 发现没有

wb = load_workbook("%s%s.xlsx"%(direction, xlsx_name))


for sheet_n in wb:
    li_sh.append(sheet_n)


def edit(sheet, d_name=do_name):
    st_name = []
    # sheet.insert_cols(3,1)

    for each_cell in sheet['B']:
        if each_cell.value is not None:
            st_name.append(each_cell.value)

    for j in range(1, len(st_name)+1):
        sheet.cell(j, 23).value = 0

    for i in range(len(name)):
        for j in range(1, len(st_name)+1):
            if st_name[j-1] == name[i]:
                sheet.cell(j, 23).value += num[i]
                d_name.append(name[i])

    sheet.cell(1, 23).value = "success"
    # print(st_name)


for sheet_s in li_sh:
    edit(sheet_s)

undo_name = [x for x in name if x not in do_name]
print("undo_list:%d" % len(undo_name))

for names in undo_name:
    print(names)
wb.save("%s%s.xlsx"%(direction, xlsx_name))

wb.close()


# if __name__ == '__main__':
